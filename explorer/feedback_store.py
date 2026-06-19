"""
Feedback storage for the INPACT platform.

Each submission from the Feedback tab is appended as a row to an Excel
workbook (``data/feedback.xlsx``). Because the platform is designed to run on
Vercel — whose serverless filesystem is **read-only and ephemeral** — the
workbook is persisted by committing it back to the GitHub repository through
the GitHub Contents API. The repo therefore doubles as the durable store, and
the resulting ``.xlsx`` can be downloaded straight from GitHub.

Behaviour is driven entirely by environment variables, so the same code path
works locally and in production:

    GITHUB_TOKEN   A fine-grained / classic PAT with `contents: write` on the
                   target repo. If unset, GitHub syncing is skipped and the
                   workbook is written to local disk instead (handy for local
                   development — just open data/feedback.xlsx).
    GITHUB_REPO    "owner/name" of the repo to write into.
                   Defaults to "MehdiGalway/INPACT_V2".
    GITHUB_BRANCH  Branch to commit to. Defaults to "main".
    FEEDBACK_PATH  Path of the workbook inside the repo.
                   Defaults to "data/feedback.xlsx".

Nothing here raises out to the request handler: every public function returns a
small status dict so the view can respond gracefully whether or not GitHub is
configured.
"""
from __future__ import annotations

import base64
import io
import os
import threading
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook

# explorer/feedback_store.py -> explorer -> <project root>
BASE_DIR = Path(__file__).resolve().parent.parent

# Column order for the workbook. Keep in sync with the front-end form fields.
COLUMNS = [
    "timestamp_utc",
    "policy_reduced_emissions",   # Yes / No / Unsure
    "effectiveness_rating",       # 1-10
    "most_affected_sector",
    "respondent_role",
    "comments",
]

# A module-level lock keeps concurrent submissions from corrupting the workbook
# while it is being read-modified-written within a single process.
_LOCK = threading.Lock()


# ---------------------------------------------------------------------------
# Configuration helpers
# ---------------------------------------------------------------------------
def _cfg() -> dict:
    return {
        "token": os.environ.get("GITHUB_TOKEN", "").strip(),
        "repo": os.environ.get("GITHUB_REPO", "MehdiGalway/INPACT_V2").strip(),
        "branch": os.environ.get("GITHUB_BRANCH", "main").strip(),
        "path": os.environ.get("FEEDBACK_PATH", "data/feedback.xlsx").strip(),
    }


def _local_path(cfg: dict) -> Path:
    return BASE_DIR / cfg["path"]


# ---------------------------------------------------------------------------
# Workbook construction
# ---------------------------------------------------------------------------
def _new_workbook_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Feedback"
    ws.append(COLUMNS)
    # Light styling for the header row.
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    _autosize(ws)
    return _to_bytes(wb)


def _to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _autosize(ws) -> None:
    widths = {}
    for row in ws.iter_rows(values_only=True):
        for i, value in enumerate(row):
            length = len(str(value)) if value is not None else 0
            widths[i] = max(widths.get(i, 10), min(length + 2, 60))
    for i, width in widths.items():
        ws.column_dimensions[chr(65 + i)].width = width


def _append_row(existing: bytes | None, row: list) -> bytes:
    """Return new workbook bytes with ``row`` appended to the Feedback sheet."""
    if existing:
        wb = load_workbook(io.BytesIO(existing))
        ws = wb["Feedback"] if "Feedback" in wb.sheetnames else wb.active
    else:
        wb = load_workbook(io.BytesIO(_new_workbook_bytes()))
        ws = wb["Feedback"]
    ws.append(row)
    _autosize(ws)
    return _to_bytes(wb)


def _row_from_payload(data: dict) -> list:
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    rating = data.get("effectiveness_rating", "")
    try:
        rating = int(rating)
    except (TypeError, ValueError):
        rating = ""
    return [
        ts,
        str(data.get("policy_reduced_emissions", "")).strip()[:20],
        rating,
        str(data.get("most_affected_sector", "")).strip()[:60],
        str(data.get("respondent_role", "")).strip()[:80],
        str(data.get("comments", "")).strip()[:1000],
    ]


# ---------------------------------------------------------------------------
# GitHub Contents API
# ---------------------------------------------------------------------------
def _gh_headers(token: str) -> dict:
    return {
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
    }


def _gh_url(cfg: dict) -> str:
    return f"https://api.github.com/repos/{cfg['repo']}/contents/{cfg['path']}"


def _gh_get_file(cfg: dict):
    """Return (content_bytes, sha) for the workbook, or (None, None) if absent."""
    import requests

    resp = requests.get(
        _gh_url(cfg),
        headers=_gh_headers(cfg["token"]),
        params={"ref": cfg["branch"]},
        timeout=15,
    )
    if resp.status_code == 200:
        body = resp.json()
        content = base64.b64decode(body["content"]) if body.get("content") else None
        return content, body.get("sha")
    if resp.status_code == 404:
        return None, None
    resp.raise_for_status()
    return None, None


def _gh_put_file(cfg: dict, content: bytes, sha: str | None) -> None:
    import requests

    payload = {
        "message": "Add feedback submission",
        "content": base64.b64encode(content).decode("ascii"),
        "branch": cfg["branch"],
    }
    if sha:
        payload["sha"] = sha
    resp = requests.put(
        _gh_url(cfg),
        headers=_gh_headers(cfg["token"]),
        json=payload,
        timeout=20,
    )
    resp.raise_for_status()


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------
def save_feedback(data: dict) -> dict:
    """
    Persist a single feedback submission.

    Returns a status dict: ``{"ok": bool, "storage": "github"|"local",
    "detail": str}``. Never raises — storage failures are reported in the dict
    so the request can still complete.
    """
    cfg = _cfg()
    row = _row_from_payload(data)

    with _LOCK:
        if cfg["token"]:
            try:
                existing, sha = _gh_get_file(cfg)
                updated = _append_row(existing, row)
                _gh_put_file(cfg, updated, sha)
                return {"ok": True, "storage": "github",
                        "detail": f"Committed to {cfg['repo']}/{cfg['path']}"}
            except Exception as exc:  # noqa: BLE001 - report, don't crash
                # Fall through to local so the submission is not lost in dev.
                local_result = _save_local(cfg, row)
                local_result["detail"] = (
                    f"GitHub sync failed ({exc.__class__.__name__}); "
                    f"saved locally instead."
                )
                local_result["ok"] = True
                return local_result
        return _save_local(cfg, row)


def _save_local(cfg: dict, row: list) -> dict:
    path = _local_path(cfg)
    path.parent.mkdir(parents=True, exist_ok=True)
    existing = path.read_bytes() if path.exists() else None
    updated = _append_row(existing, row)
    try:
        path.write_bytes(updated)
        return {"ok": True, "storage": "local",
                "detail": f"Saved to {path} (set GITHUB_TOKEN to sync to GitHub)."}
    except OSError as exc:
        # Read-only filesystem (e.g. Vercel) with no token configured.
        return {"ok": False, "storage": "none",
                "detail": f"Could not persist feedback: {exc}. "
                          f"Configure GITHUB_TOKEN to enable durable storage."}


def get_workbook_bytes() -> bytes | None:
    """
    Return the current workbook bytes for download, preferring GitHub (the
    source of truth in production) and falling back to local disk. Returns
    ``None`` if no workbook exists yet.
    """
    cfg = _cfg()
    if cfg["token"]:
        try:
            content, _sha = _gh_get_file(cfg)
            if content is not None:
                return content
        except Exception:  # noqa: BLE001 - fall back to local copy
            pass
    path = _local_path(cfg)
    if path.exists():
        return path.read_bytes()
    return None
